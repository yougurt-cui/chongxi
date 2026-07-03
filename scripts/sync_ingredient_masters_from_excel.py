#!/usr/bin/env python3
"""Synchronize standard ingredients and aliases from the corrected workbook."""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import pymysql


BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from app_config import get_mysql_config  # noqa: E402


STANDARD_TABLE = "catfood_standard_ingredient"
ALIAS_TABLE = "catfood_standard_ingredient_alias"


def _text(value: Any) -> str:
    return str(value or "").strip()


def _load(path: Path) -> tuple[dict[str, str], list[dict[str, Any]], dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    standard_rows = list(workbook["标准名汇总"].iter_rows(min_row=2, values_only=True))
    alias_rows = list(workbook["别名映射表"].iter_rows(min_row=2, values_only=True))

    standards: dict[str, str] = {}
    repeated_ids: dict[str, list[str]] = defaultdict(list)
    for raw_id, raw_name, *_ in standard_rows:
        standard_id, standard_name = _text(raw_id), _text(raw_name)
        if not standard_id and not standard_name:
            continue
        if not standard_id or not standard_name:
            raise ValueError(f"标准名汇总存在不完整行: {(raw_id, raw_name)}")
        repeated_ids[standard_id].append(standard_name)
        standards[standard_id] = standard_name  # The last corrected row wins.

    aliases = []
    for raw_alias_id, raw_standard_id, raw_alias_name, raw_normalized, *_ in alias_rows:
        if all(value is None or not _text(value) for value in (raw_alias_id, raw_standard_id, raw_alias_name, raw_normalized)):
            continue
        alias_id = int(raw_alias_id)
        standard_id = _text(raw_standard_id)
        alias_name = _text(raw_alias_name)
        normalized_alias = _text(raw_normalized)
        if not standard_id or not alias_name or not normalized_alias:
            raise ValueError(f"别名映射表存在不完整行: {(raw_alias_id, raw_standard_id, raw_alias_name, raw_normalized)}")
        if standard_id not in standards:
            raise ValueError(f"别名 {alias_id} 引用了不存在的标准 ID: {standard_id}")
        aliases.append(
            {
                "alias_id": alias_id,
                "standard_ingredient_id": standard_id,
                "alias_name": alias_name,
                "normalized_alias": normalized_alias,
            }
        )

    alias_ids = [row["alias_id"] for row in aliases]
    if len(alias_ids) != len(set(alias_ids)):
        raise ValueError("别名映射表存在重复 alias_id")

    name_counts = Counter(standards.values())
    alias_name_targets: dict[str, set[str]] = defaultdict(set)
    for row in aliases:
        alias_name_targets[row["alias_name"]].add(row["standard_ingredient_id"])
    diagnostics = {
        "duplicate_standard_ids": {
            key: values for key, values in repeated_ids.items() if len(values) > 1
        },
        "duplicate_standard_names": {
            key: count for key, count in name_counts.items() if count > 1
        },
        "ambiguous_alias_names": {
            key: sorted(values) for key, values in alias_name_targets.items() if len(values) > 1
        },
    }
    return standards, aliases, diagnostics


def _connect():
    return pymysql.connect(
        **get_mysql_config(), cursorclass=pymysql.cursors.DictCursor, autocommit=False
    )


def _backup(cursor, table: str, suffix: str) -> str:
    backup = f"{table}_bak_{suffix}"
    cursor.execute(f"CREATE TABLE `{backup}` LIKE `{table}`")
    cursor.execute(f"INSERT INTO `{backup}` SELECT * FROM `{table}`")
    return backup


def _index_names(cursor, table: str) -> dict[str, dict[str, Any]]:
    cursor.execute(f"SHOW INDEX FROM `{table}`")
    return {row["Key_name"]: row for row in cursor.fetchall()}


def sync(path: Path, *, apply: bool) -> dict[str, Any]:
    standards, aliases, diagnostics = _load(path)
    with _connect() as conn:
        with conn.cursor() as cursor:
            cursor.execute(f"SELECT standard_ingredient_id, standard_name, active FROM `{STANDARD_TABLE}`")
            current_standards = {row["standard_ingredient_id"]: row for row in cursor.fetchall()}
            cursor.execute(f"SELECT alias_id, standard_ingredient_id, alias_name, normalized_alias FROM `{ALIAS_TABLE}`")
            current_aliases = list(cursor.fetchall())

            added_ids = sorted(set(standards) - set(current_standards))
            deactivated_ids = sorted(set(current_standards) - set(standards))
            renamed = [
                {
                    "standard_ingredient_id": standard_id,
                    "before": current_standards[standard_id]["standard_name"],
                    "after": standard_name,
                }
                for standard_id, standard_name in standards.items()
                if standard_id in current_standards
                and current_standards[standard_id]["standard_name"] != standard_name
            ]

            backups = []
            if apply:
                suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
                backups = [
                    _backup(cursor, STANDARD_TABLE, suffix),
                    _backup(cursor, ALIAS_TABLE, suffix),
                ]

                standard_indexes = _index_names(cursor, STANDARD_TABLE)
                if "uq_standard_ingredient_name" in standard_indexes:
                    cursor.execute(
                        f"ALTER TABLE `{STANDARD_TABLE}` DROP INDEX `uq_standard_ingredient_name`"
                    )
                standard_indexes = _index_names(cursor, STANDARD_TABLE)
                if "idx_standard_ingredient_name" not in standard_indexes:
                    cursor.execute(
                        f"ALTER TABLE `{STANDARD_TABLE}` ADD INDEX `idx_standard_ingredient_name` (`standard_name`)"
                    )

                alias_indexes = _index_names(cursor, ALIAS_TABLE)
                if "uq_ingredient_alias_normalized" in alias_indexes:
                    cursor.execute(
                        f"ALTER TABLE `{ALIAS_TABLE}` DROP INDEX `uq_ingredient_alias_normalized`"
                    )
                alias_indexes = _index_names(cursor, ALIAS_TABLE)
                if "idx_ingredient_alias_normalized" not in alias_indexes:
                    cursor.execute(
                        f"ALTER TABLE `{ALIAS_TABLE}` ADD INDEX `idx_ingredient_alias_normalized` (`normalized_alias`)"
                    )
                if "idx_ingredient_alias_name" not in alias_indexes:
                    cursor.execute(
                        f"ALTER TABLE `{ALIAS_TABLE}` ADD INDEX `idx_ingredient_alias_name` (`alias_name`)"
                    )

                cursor.execute(f"UPDATE `{STANDARD_TABLE}` SET active=0")
                for standard_id, standard_name in standards.items():
                    cursor.execute(
                        f"""
                        INSERT INTO `{STANDARD_TABLE}`(standard_ingredient_id, standard_name, active)
                        VALUES(%s, %s, 1)
                        ON DUPLICATE KEY UPDATE standard_name=VALUES(standard_name), active=1
                        """,
                        (standard_id, standard_name),
                    )

                cursor.execute(f"DELETE FROM `{ALIAS_TABLE}`")
                cursor.executemany(
                    f"""
                    INSERT INTO `{ALIAS_TABLE}`(
                      alias_id, standard_ingredient_id, alias_name, normalized_alias,
                      source, confidence, active
                    ) VALUES(%s, %s, %s, %s, 'corrected_excel', 1.00000, 1)
                    """,
                    [
                        (
                            row["alias_id"],
                            row["standard_ingredient_id"],
                            row["alias_name"],
                            row["normalized_alias"],
                        )
                        for row in aliases
                    ],
                )
                conn.commit()
            else:
                conn.rollback()

    return {
        "applied": apply,
        "source": str(path),
        "backups": backups,
        "standards_in_excel": len(standards),
        "aliases_in_excel": len(aliases),
        "standards_before": len(current_standards),
        "aliases_before": len(current_aliases),
        "standard_ids_added": added_ids,
        "standard_ids_deactivated": deactivated_ids,
        "standards_renamed": renamed,
        "diagnostics": diagnostics,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("excel", type=Path)
    parser.add_argument("--apply", action="store_true", help="commit changes; default previews")
    parser.add_argument("--report", type=Path)
    args = parser.parse_args()
    result = sync(args.excel.expanduser().resolve(), apply=bool(args.apply))
    output = json.dumps(result, ensure_ascii=False, default=str, indent=2)
    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(output + "\n", encoding="utf-8")
    print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
