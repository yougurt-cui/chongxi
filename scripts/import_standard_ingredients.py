#!/usr/bin/env python3
"""Import standard ingredient master data from the shared Excel file."""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl
import pymysql


BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from app_config import get_mysql_config  # noqa: E402


DEFAULT_EXCEL_PATH = Path("/Users/yoghourt/原材料标准化表.xlsx")
INGREDIENT_TABLE = "catfood_standard_ingredient"
INGREDIENT_ALIAS_TABLE = "catfood_standard_ingredient_alias"


def normalize_ingredient_key(value: Any) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[\s·•._\-—–/\\|,:：;；，。()（）\[\]【】'\"®™]+", "", text)


def _clean(value: Any) -> str:
    text = str(value or "").strip()
    return "" if text.lower() == "nan" else re.sub(r"\s+", " ", text)


def _connect():
    return pymysql.connect(
        **get_mysql_config(),
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=False,
    )


def ensure_tables(cursor) -> None:
    cursor.execute(
        f"""
        CREATE TABLE IF NOT EXISTS `{INGREDIENT_TABLE}` (
          standard_ingredient_id VARCHAR(64) NOT NULL,
          standard_name VARCHAR(255) NOT NULL,
          ingredient_family VARCHAR(128) NULL,
          source_type VARCHAR(64) NULL,
          animal_source VARCHAR(128) NULL,
          primary_nutrition_role VARCHAR(128) NULL,
          active TINYINT NOT NULL DEFAULT 1,
          created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
          updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          PRIMARY KEY (standard_ingredient_id),
          UNIQUE KEY uq_standard_ingredient_name (standard_name),
          KEY idx_ingredient_family (ingredient_family),
          KEY idx_source_type (source_type)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """
    )
    cursor.execute(
        f"""
        CREATE TABLE IF NOT EXISTS `{INGREDIENT_ALIAS_TABLE}` (
          alias_id BIGINT UNSIGNED NOT NULL AUTO_INCREMENT,
          standard_ingredient_id VARCHAR(64) NOT NULL,
          alias_name VARCHAR(255) NOT NULL,
          normalized_alias VARCHAR(255) NOT NULL,
          source VARCHAR(64) NULL,
          confidence DECIMAL(6,5) NOT NULL DEFAULT 1.00000,
          active TINYINT NOT NULL DEFAULT 1,
          created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
          PRIMARY KEY (alias_id),
          UNIQUE KEY uq_ingredient_alias_normalized (normalized_alias),
          KEY idx_ingredient_alias_standard_id (standard_ingredient_id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """
    )


def read_excel_rows(path: Path) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    header = [str(value or "").strip() for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    required = {
        "standard_ingredient_id",
        "standard_name",
        "ingredient_family",
        "source_type",
        "animal_source",
        "primary_nutrition_role",
    }
    missing = required - set(header)
    if missing:
        raise ValueError(f"Excel 缺少字段: {sorted(missing)}")
    rows: list[dict[str, str]] = []
    for raw_row in sheet.iter_rows(min_row=2, values_only=True):
        item = {header[index]: _clean(value) for index, value in enumerate(raw_row)}
        if item.get("standard_ingredient_id") and item.get("standard_name"):
            rows.append(item)
    return rows


def import_standard_ingredients(path: Path, *, apply: bool) -> dict[str, Any]:
    rows = read_excel_rows(path)
    with _connect() as conn:
        with conn.cursor() as cursor:
            ensure_tables(cursor)
            for row in rows:
                cursor.execute(
                    f"""
                    INSERT INTO `{INGREDIENT_TABLE}`(
                      standard_ingredient_id, standard_name, ingredient_family,
                      source_type, animal_source, primary_nutrition_role, active
                    ) VALUES(%s, %s, %s, %s, %s, %s, 1)
                    ON DUPLICATE KEY UPDATE
                      standard_name = VALUES(standard_name),
                      ingredient_family = VALUES(ingredient_family),
                      source_type = VALUES(source_type),
                      animal_source = VALUES(animal_source),
                      primary_nutrition_role = VALUES(primary_nutrition_role),
                      active = 1
                    """,
                    (
                        row["standard_ingredient_id"],
                        row["standard_name"],
                        row.get("ingredient_family") or None,
                        row.get("source_type") or None,
                        row.get("animal_source") or None,
                        row.get("primary_nutrition_role") or None,
                    ),
                )
                cursor.execute(
                    f"""
                    INSERT INTO `{INGREDIENT_ALIAS_TABLE}`(
                      standard_ingredient_id, alias_name, normalized_alias,
                      source, confidence, active
                    ) VALUES(%s, %s, %s, 'standard_ingredient_excel', 1.0, 1)
                    ON DUPLICATE KEY UPDATE
                      standard_ingredient_id = VALUES(standard_ingredient_id),
                      alias_name = VALUES(alias_name),
                      source = VALUES(source),
                      confidence = VALUES(confidence),
                      active = 1
                    """,
                    (
                        row["standard_ingredient_id"],
                        row["standard_name"],
                        normalize_ingredient_key(row["standard_name"]),
                    ),
                )
        if apply:
            conn.commit()
        else:
            conn.rollback()
    return {"ok": True, "applied": apply, "excel_path": str(path), "row_count": len(rows)}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import cat-food standard ingredient master Excel.")
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL_PATH), help="standard ingredient Excel path")
    parser.add_argument("--apply", action="store_true", help="write changes; default is dry-run rollback")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    result = import_standard_ingredients(Path(args.excel).expanduser(), apply=bool(args.apply))
    print(result)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
