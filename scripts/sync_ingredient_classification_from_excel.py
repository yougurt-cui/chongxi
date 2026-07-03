#!/usr/bin/env python3
"""Synchronize standard ingredient classification fields from Excel."""

from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import pymysql


BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from app_config import get_mysql_config  # noqa: E402


TABLE = "catfood_standard_ingredient"
SOURCE_TYPE_MAP = {
    "动物来源": "animal",
    "植物来源": "plant",
    "矿物来源": "mineral",
    "微生物来源": "microbial",
    "合成/纯化来源": "synthetic",
    "复合来源": "mixed",
}


def _text(value: Any) -> str:
    return str(value or "").strip()


def _connect():
    return pymysql.connect(
        **get_mysql_config(), cursorclass=pymysql.cursors.DictCursor, autocommit=False
    )


def _load_rows(path: Path) -> dict[str, list[dict[str, str]]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [_text(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    required = {
        "standard_ingredient_id",
        "normalized_alias",
        "source_type",
        "ingredient_family",
        "primary_nutrition_role",
    }
    missing = required - set(headers)
    if missing:
        raise ValueError(f"Excel 缺少字段: {sorted(missing)}")
    positions = {name: headers.index(name) for name in required}
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and _text(value) for value in values):
            continue
        row = {name: _text(values[index]) for name, index in positions.items()}
        if not all(row.values()):
            raise ValueError(f"Excel 存在不完整分类行: {row}")
        if row["source_type"] not in SOURCE_TYPE_MAP:
            raise ValueError(f"未知 source_type: {row['source_type']}")
        grouped[row["standard_ingredient_id"]].append(row)
    return grouped


def sync(path: Path, *, apply: bool) -> dict[str, Any]:
    grouped = _load_rows(path)
    with _connect() as conn:
        with conn.cursor() as cursor:
            cursor.execute(f"SELECT * FROM `{TABLE}`")
            standards = {row["standard_ingredient_id"]: row for row in cursor.fetchall()}

            unknown_ids = sorted(set(grouped) - set(standards))
            if unknown_ids:
                raise ValueError(f"Excel 存在数据库中没有的标准 ID: {unknown_ids}")

            selected: dict[str, dict[str, str]] = {}
            duplicate_resolution = []
            for standard_id, rows in grouped.items():
                standard_name = _text(standards[standard_id]["standard_name"])
                matching = [row for row in rows if row["normalized_alias"] == standard_name]
                chosen = matching[-1] if matching else rows[-1]
                selected[standard_id] = chosen
                if len(rows) > 1:
                    duplicate_resolution.append(
                        {
                            "standard_ingredient_id": standard_id,
                            "database_standard_name": standard_name,
                            "available_names": [row["normalized_alias"] for row in rows],
                            "selected_name": chosen["normalized_alias"],
                        }
                    )

            missing_active_ids = sorted(
                standard_id
                for standard_id, row in standards.items()
                if int(row.get("active") or 0) == 1 and standard_id not in selected
            )
            changes = []
            for standard_id, row in selected.items():
                before = standards[standard_id]
                after = {
                    "source_type": SOURCE_TYPE_MAP[row["source_type"]],
                    "ingredient_family": row["ingredient_family"],
                    "primary_nutrition_role": row["primary_nutrition_role"],
                }
                if any(_text(before.get(key)) != value for key, value in after.items()):
                    changes.append(
                        {
                            "standard_ingredient_id": standard_id,
                            "standard_name": before["standard_name"],
                            "before": {key: before.get(key) for key in after},
                            "after": after,
                        }
                    )

            backup = None
            if apply:
                suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup = f"{TABLE}_bak_{suffix}"
                cursor.execute(f"CREATE TABLE `{backup}` LIKE `{TABLE}`")
                cursor.execute(f"INSERT INTO `{backup}` SELECT * FROM `{TABLE}`")
                cursor.executemany(
                    f"""
                    UPDATE `{TABLE}`
                    SET source_type=%s, ingredient_family=%s, primary_nutrition_role=%s
                    WHERE standard_ingredient_id=%s
                    """,
                    [
                        (
                            SOURCE_TYPE_MAP[row["source_type"]],
                            row["ingredient_family"],
                            row["primary_nutrition_role"],
                            standard_id,
                        )
                        for standard_id, row in selected.items()
                    ],
                )
                conn.commit()
            else:
                conn.rollback()

    return {
        "applied": apply,
        "source": str(path),
        "backup": backup,
        "excel_rows": sum(len(rows) for rows in grouped.values()),
        "classified_standard_ids": len(selected),
        "changed_standard_ids": len(changes),
        "missing_active_ids": missing_active_ids,
        "duplicate_resolution": duplicate_resolution,
        "changes": changes,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("excel", type=Path)
    parser.add_argument("--apply", action="store_true", help="commit changes; default previews")
    parser.add_argument("--report", type=Path)
    args = parser.parse_args()
    result = sync(args.excel.expanduser().resolve(), apply=bool(args.apply))
    rendered = json.dumps(result, ensure_ascii=False, default=str, indent=2)
    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(rendered + "\n", encoding="utf-8")
    summary = {key: value for key, value in result.items() if key != "changes"}
    print(json.dumps(summary, ensure_ascii=False, default=str, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
